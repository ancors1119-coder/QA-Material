# -*- coding: utf-8 -*-
"""원료 폴더의 엑셀 문서를 텍스트로 뽑아낸다.

정독은 PDF 만 훑어 왔는데, 알레르겐 목록·금지성분 대조표처럼
**판정에 직접 필요한 자료가 엑셀로 관리되는 경우**가 있다(611410 로즈마리오일).
PDF 추출본에는 아예 나타나지 않아 조용히 빠지므로 별도 도구를 둔다.

사용법
    python xls_dump.py <파일 또는 폴더> [--max 80]

    폴더를 주면 하위의 .xlsx/.xlsm/.xls 를 모두 훑는다.
    --max 는 시트당 출력할 행 수(기본 80).

빈 행과 빈 열은 접어서 출력한다 — 서식용 빈칸이 많아 그대로 찍으면 읽기 어렵다.
"""
import io, os, re, sys, glob

sys.stdout.reconfigure(encoding='utf-8')


def cells_to_lines(rows, max_rows):
    """행 목록을 사람이 읽을 수 있는 줄로 바꾼다. 값이 하나도 없는 행은 버린다."""
    out = []
    for r in rows:
        vals = ['' if c is None else str(c).strip() for c in r]
        if not any(vals):
            continue
        out.append(' | '.join(v for v in vals if v))
        if len(out) >= max_rows:
            out.append('… (이하 생략)')
            break
    return out


def dump_xlsx(path, max_rows):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        print('  [시트] %s (%s행 × %s열)' % (ws.title, ws.max_row, ws.max_column))
        for line in cells_to_lines(ws.iter_rows(values_only=True), max_rows):
            print('    ' + line)
    wb.close()


def dump_xls(path, max_rows):
    try:
        import xlrd
    except ImportError:
        print('  (.xls 는 xlrd 가 필요합니다 — pip install xlrd)')
        return
    wb = xlrd.open_workbook(path)
    for ws in wb.sheets():
        print('  [시트] %s (%d행 × %d열)' % (ws.name, ws.nrows, ws.ncols))
        rows = (ws.row_values(i) for i in range(ws.nrows))
        for line in cells_to_lines(rows, max_rows):
            print('    ' + line)


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    src = sys.argv[1]
    max_rows = 80
    if '--max' in sys.argv:
        max_rows = int(sys.argv[sys.argv.index('--max') + 1])

    if os.path.isdir(src):
        paths = []
        for ext in ('xlsx', 'xlsm', 'xls', 'XLSX', 'XLS'):
            paths += glob.glob(os.path.join(src, '**', '*.' + ext), recursive=True)
        paths = sorted(set(paths))
    else:
        paths = [src]

    for p in paths:
        print('=' * 76)
        print('■', os.path.basename(p))
        try:
            if p.lower().endswith('.xls'):
                dump_xls(p, max_rows)
            else:
                dump_xlsx(p, max_rows)
        except Exception as e:
            print('  (읽기 실패: %s)' % e)
    print('\n엑셀 %d건 처리' % len(paths))
    return 0


def demo():
    """자체 점검 — 임시 엑셀을 만들어 빈 행이 접히는지 본다."""
    import tempfile, openpyxl
    with tempfile.TemporaryDirectory() as d:
        p = os.path.join(d, 't.xlsx')
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'ALLERGEN'
        ws.append(['CAS', 'Name', '%'])
        ws.append([None, None, None])              # 빈 행 — 접혀야 한다
        ws.append(['5989-27-5', 'D-Limonene', 6.0])
        wb.save(p); wb.close()

        buf = io.StringIO(); old = sys.stdout; sys.stdout = buf
        try:
            dump_xlsx(p, 80)
        finally:
            sys.stdout = old
        txt = buf.getvalue()
    assert 'D-Limonene' in txt, '값이 출력되지 않았습니다'
    assert txt.count('|') == 4, '빈 행이 접히지 않았습니다: %r' % txt
    print('xls_dump 자체 점검 통과')


if __name__ == '__main__':
    if len(sys.argv) == 2 and sys.argv[1] == '--demo':
        demo()
    else:
        sys.exit(main())
